# -*- coding: utf-8 -*-
"""
Exportacao da Analise de Margem (paleta institucional N1, fonte Arial).

Duas abas:
  - "Análise de Margem"  : priorizados + pedidos limpos
  - "Pedidos em Atenção" : nao priorizados sem rota / MC negativa / bloqueados

Destaques de linha:
  - MC negativa -> fundo vermelho claro (#FAECE7) e MC$/MC% em vermelho
  - FOB         -> fundo verde-claro (#E8F5F0) e celula 'Tipo Frete' destacada
  - zebra       -> #F8F9FA nas demais
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERDE_PRIMARIO = "008D67"
BRANCO = "FFFFFF"
INK = "333333"
CINZA_GELO = "E9ECEF"
ZEBRA = "F8F9FA"
RISCO = "DC3545"
RISCO_FILL = "FAECE7"
FOB_FILL = "E8F5F0"
FOB_TXT = "006B4F"
FONTE = "Arial"

FMT_MOEDA = u'"R$" #,##0.00'
FMT_PCT = '0.00%'

_thin = Side(style="thin", color=CINZA_GELO)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_wthin = Side(style="thin", color=BRANCO)
_HBORDER = Border(left=_wthin, right=_wthin, top=_wthin, bottom=_wthin)


def _write_sheet(ws, df, meta, fob_highlight=True):
    cols = list(df.columns)
    money = set(meta.get("money_cols", []))
    pct = set(meta.get("pct_cols", []))
    mc_nom = "Margem de Contribuição Nominal (MC$)"
    mc_pct = "Margem de Contribuição Percentual (MC%)"
    tf_col = "Tipo Frete"

    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor=VERDE_PRIMARIO)
    header_font = Font(color=BRANCO, bold=True, size=9, name=FONTE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    neg_fill = PatternFill("solid", fgColor=RISCO_FILL)
    fob_fill = PatternFill("solid", fgColor=FOB_FILL)
    zebra_fill = PatternFill("solid", fgColor=ZEBRA)

    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=str(col))
        c.fill = header_fill; c.font = header_font
        c.alignment = header_align; c.border = _HBORDER
    ws.freeze_panes = "A2"

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        try:
            mcv = float(row[mc_nom])
        except Exception:
            mcv = None
        negativo = (mcv is not None and mcv < 0)
        is_fob = str(row.get(tf_col, "")).strip().upper() == "FOB"
        for j, col in enumerate(cols, start=1):
            val = row[col]
            if col in money or col in pct:
                try:
                    val = float(val)
                except Exception:
                    val = 0.0
            c = ws.cell(row=i, column=j, value=val)
            c.border = _BORDER
            color, bold = INK, False
            if negativo and col in (mc_nom, mc_pct):
                color, bold = RISCO, True
            if fob_highlight and is_fob and col == tf_col:
                color, bold = FOB_TXT, True
            c.font = Font(name=FONTE, size=9, color=color, bold=bold)
            if col in money:
                c.number_format = FMT_MOEDA
            elif col in pct:
                c.number_format = FMT_PCT
            if negativo:
                c.fill = neg_fill
            elif fob_highlight and is_fob:
                c.fill = fob_fill
            elif i % 2 == 1:
                c.fill = zebra_fill

    for j, col in enumerate(cols, start=1):
        letter = get_column_letter(j)
        max_len = len(str(col))
        for i in range(2, ws.max_row + 1):
            v = ws.cell(row=i, column=j).value
            if v is None:
                continue
            txt = f"{v:,.2f}" if isinstance(v, float) else str(v)
            max_len = max(max_len, len(txt))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 10), 48)
    ws.row_dimensions[1].height = 30


def _write_consolidado(ws, df):
    ws.sheet_view.showGridLines = False
    cols = list(df.columns)
    money = {"Faturamento (R$)", "MC (R$)"}
    pctc = {"MC (%)"}
    numc = {"Volume (t)"}
    header_fill = PatternFill("solid", fgColor=VERDE_PRIMARIO)
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=str(col))
        c.fill = header_fill; c.font = Font(color=BRANCO, bold=True, size=10, name=FONTE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _HBORDER
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        total = str(row[cols[0]]).startswith("TOTAL")
        mcv = float(row["MC (R$)"]) if "MC (R$)" in df.columns else 0.0
        for j, col in enumerate(cols, start=1):
            val = row[col]
            if col in money or col in pctc or col in numc:
                try: val = float(val)
                except Exception: val = 0.0
            c = ws.cell(row=i, column=j, value=val)
            c.border = _BORDER
            color = RISCO if (mcv < 0 and col in ("MC (R$)", "MC (%)")) else INK
            c.font = Font(name=FONTE, size=10, color=color, bold=total)
            if col in money: c.number_format = FMT_MOEDA
            elif col in pctc: c.number_format = FMT_PCT
            elif col in numc: c.number_format = '#,##0.00'
            if total: c.fill = PatternFill("solid", fgColor=ZEBRA)
    widths = [30, 14, 20, 20, 12]
    for j, w in enumerate(widths[:len(cols)], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _empty_sheet(ws, msg, width=70):
    ws.sheet_view.showGridLines = False
    c = ws.cell(row=1, column=1, value=msg)
    c.font = Font(name=FONTE, size=10, color=INK, italic=True)
    ws.column_dimensions["A"].width = width


def export_excel(df_main, df_exc, df_fob, meta, out_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Análise de Margem"      # todos os pedidos (revenda com custo exato)
    if df_main is not None and not df_main.empty:
        _write_sheet(ws1, df_main, meta)
    else:
        _empty_sheet(ws1, "Nenhum pedido nesta carteira.")

    ws2 = wb.create_sheet("Pedidos em Atenção")
    if df_exc is not None and not df_exc.empty:
        _write_sheet(ws2, df_exc, meta)
    else:
        _empty_sheet(ws2, "Nenhum pedido em atenção (sem rota / MC negativa / bloqueado não priorizado).", 80)

    ws3 = wb.create_sheet("Pedidos FOB - Retirada")
    if df_fob is not None and not df_fob.empty:
        _write_sheet(ws3, df_fob, meta, fob_highlight=False)
        ws3.column_dimensions["A"].width = 24  # coluna manual de data
    else:
        _empty_sheet(ws3, "Nenhum pedido FOB nesta carteira.", 60)

    wb.save(out_path)
    return out_path
